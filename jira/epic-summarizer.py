#!/usr/bin/env python3
# INSTALLATION
# ------------
# pip install jira
# pip install openpyxl
# pip install openpyxl-stubs
# pip install fuzzywuzzy
# pip install python-Levenshtein
# pip install wordcloud
# RUN
# ---
# python epic-summarizer.py

from jira import JIRA
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.comments import Comment
from wordcloud import WordCloud, STOPWORDS
from fuzzywuzzy import fuzz
from statistics import mean
import math
import os
import sys
import json

CF_CUSTOM = {"name": "Custom Field", "id": "customfield_10345", "type": "array"}
CF_FITMENT = {"name": "Fitment Level", "id": "customfield_10362", "type": "option"}
CF_SIZING = {"name": "Sizing", "id": "customfield_10357", "type": "option"}
# Assuming size 2.5 for Not Yet Sized
SIZING_MATRIX = {
    "No Data": -1,
    None: 2.5,
    "N/A": 0,
    "X-Small": 1,
    "Small": 2,
    "Medium": 3,
    "Large": 4,
    "X-Large": 5,
    "XX-Large": 6,
}
SIZING_REVERSE_MATRIX = {v: k for k, v in SIZING_MATRIX.items()}
STOPWORDS = {
    "https",
    "atlassian",
    "lucid",
    "fitment",
    "lucidchart",
    "invitationId",
    "Size",
    "pdf",
    "link",
    "net",
}


def main(args):
    JIRA_SERVER = os.environ.get("JIRA_SERVER", "https://jira.atlassian.net")
    JIRA_TOKEN = os.environ.get("JIRA_TOKEN")
    JIRA_EMAIL = os.environ.get("JIRA_EMAIL")

    if not JIRA_TOKEN:
        print("Please set the JIRA_TOKEN env variable.")
        sys.exit(1)

    # Some Authentication Methods
    jira = JIRA(server=JIRA_SERVER, basic_auth=(str(JIRA_EMAIL), JIRA_TOKEN))

    component = args[0]

    # Update this with actual Jira project name
    componentIssues = jira.search_issues(
        f'project = JIRA_PROJECT_NAME AND component = "{component}"'
    )

    for issue in componentIssues:
        analyzeIssue(jira, component, issue)


def analyzeIssue(jira: JIRA, component, issue):
    # create component folder
    if not os.path.isdir(f"generated/{component}"):
        os.makedirs(f"generated/{component}")
    linkedIssues = getLinkedIssues(jira, issue)
    liSizes = linkedIssueSizes(issue.fields.description, linkedIssues)
    proposedSize = proposeSize(liSizes)
    createIssueWorkbook(component, issue, linkedIssues, liSizes, proposedSize)


def createIssueWorkbook(component, issue, linkedIssues, liSizes, proposedSize):
    # key issue fields
    key = issue.key

    # begin:workbook
    workbook = Workbook()

    # create issue folder
    if not os.path.isdir(f"generated/{component}/{key}"):
        os.makedirs(f"generated/{component}/{key}/WC")

    # create sheets
    summarySheet(workbook, component, issue, linkedIssues, liSizes, proposedSize)
    wordCloudSheet(workbook, component, issue, linkedIssues)

    # save: workbook
    workbook.save(f"generated/{component}/{key}/{key}.xlsx")
    print(f"Wrote Sizing Sheet: generated/{component}/{key}/{key}.xlsx")


def summarySheet(workbook, component, issue, linkedIssues, liSizes, proposedSize):
    sheet = workbook.active
    sheet.title = "Summary"

    headings = [
        "Key",
        "Summary",
        "Description",
        "Proposed Sizing",
        "LI: Key",
        "LI: Type",
        "LI: Summary",
        "LI: Description",
        "LI: Customer",
        "LI: Fitment",
        "LI: Size",
        "FM: Ratio",
        "FM: Size",
    ]
    sheet.append(headings)
    summary = issue.fields.summary
    description = issue.fields.description
    issueData = [
        '=HYPERLINK("https://jira.atlassian.net/browse/{}", "{}")'.format(
            issue.key, issue.key
        ),
        summary,
        description,
        proposedSize,
    ]
    issueDataSpacer = ["", "", "", ""]
    sheet.append(issueData)
    for linkedIssue in linkedIssues:
        fm = [
            liSizes[linkedIssue[0]]["fm"],
            SIZING_REVERSE_MATRIX[liSizes[linkedIssue[0]]["fmSize"]],
        ]
        sheet.append(issueDataSpacer + linkedIssue + fm)
    # Style
    styleHeader(sheet, "A1:M1")
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 65
    sheet.column_dimensions["G"].width = 25
    sheet.column_dimensions["H"].width = 65
    sheet["D2"].style = "Neutral"
    wrapText(sheet)
    comment = Comment(json.dumps(linkedSizesDump(liSizes), indent=1), "Sizer")
    comment.width = 500
    comment.height = 500
    sheet["D2"].comment = comment


def wordCloudSheet(workbook, component, issue, linkedIssues):
    wordCloudSheet = workbook.create_sheet("Word Cloud")
    headings = ["Key", "Summary", "Relation", "Word Cloud"]
    wordCloudSheet.append(headings)
    # Issue
    wordCloudSheet.append([issue.key, issue.fields.summary, ""])
    wcStopWords = stopWords()
    cloud = WordCloud(stopwords=wcStopWords).generate(issue.fields.description)
    wcLocation = f"generated/{component}/{issue.key}/WC"
    cloud.to_file(f"{wcLocation}/{issue.key}-Wc.png")
    img = Image(f"{wcLocation}/{issue.key}-Wc.png")
    wordCloudSheet.add_image(img, "D2")
    # Linked Issues
    for idx, linkedIssue in enumerate(linkedIssues):
        cloud = WordCloud(stopwords=wcStopWords).generate(linkedIssue[3])
        cloud.to_file(f"{wcLocation}/{linkedIssue[0]}-Wc.png")
        wordCloudSheet.append([linkedIssue[0], linkedIssue[2], linkedIssue[1]])
        img = Image(f"{wcLocation}/{linkedIssue[0]}-Wc.png")
        wordCloudSheet.add_image(img, f"D{idx + 3}")
    # Style
    styleHeader(wordCloudSheet, "A1:D1")
    wordCloudSheet.column_dimensions["D"].width = 55
    for row in range(2, wordCloudSheet.max_row + 1):
        wordCloudSheet.row_dimensions[row].height = 155
    wordCloudSheet.column_dimensions["A"].width = 30
    wordCloudSheet.column_dimensions["B"].width = 60
    wordCloudSheet.column_dimensions["C"].width = 30
    wrapText(wordCloudSheet)


def linkedSizesDump(liSizes):
    rliSizes = {}
    for k, v in liSizes.items():
        rliSizes.update(
            {
                k: SIZING_REVERSE_MATRIX[v["size"]],
                v["fm"]: SIZING_REVERSE_MATRIX[v["fmSize"]],
            }
        )
    return rliSizes


def stopWords():
    return STOPWORDS | STOPWORDS


def styleHeader(sheet, range):
    for row in sheet[range]:
        for cell in row:
            cell.style = "Accent1"


def wrapText(sheet):
    wrapText = Alignment(wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = wrapText


def getLinkedIssues(jira: JIRA, issue):
    linkedIssues = []
    for issue_link in issue.fields.issuelinks:
        liKey = None
        if hasattr(issue_link, "outwardIssue"):
            outwardIssue = issue_link.outwardIssue
            liKey = outwardIssue.key
        if hasattr(issue_link, "inwardIssue"):
            inwardIssue = issue_link.inwardIssue
            liKey = inwardIssue.key
        if liKey != None:
            liType = issue_link.type.name
            if liType == "Gantt End to Start":
                linkedIssue = jira.issue(liKey)
                liSummary = linkedIssue.fields.summary
                liDescription = linkedIssue.fields.description
                liCustomField = cfValue(linkedIssue, CF_CUSTOM)
                liFitment = cfValue(linkedIssue, CF_FITMENT)
                liSizing = cfValue(linkedIssue, CF_SIZING)
                linkedIssues.append(
                    [
                        liKey,
                        liType,
                        liSummary,
                        liDescription,
                        liCustomField,
                        liFitment,
                        liSizing,
                    ]
                )
    return linkedIssues


def linkedIssueSizes(matchString, linkedIssues):
    linkedIssueSizes = {}
    for linkedIssue in linkedIssues:
        fm = fuzzyMatch(matchString, linkedIssue[3])
        size = SIZING_MATRIX[linkedIssue[6]]
        linkedIssueSizes.update(
            {
                linkedIssue[0]: {
                    "size": size,
                    "fm": fm,
                    "fmSize": math.ceil((fm / 100) * size),
                }
            }
        )
    return linkedIssueSizes


def proposeSize(linkedIssueSizes):
    if linkedIssueSizes:
        liSizes = linkedIssueSizes.values()
        fmSizes = [d["fmSize"] for d in liSizes]
        umSizes = [d["size"] - d["fmSize"] for d in liSizes]
        meanUnmSize = math.ceil(mean(umSizes))
        meanFmSize = math.ceil(mean(fmSizes))
        return SIZING_REVERSE_MATRIX[meanFmSize + meanUnmSize]
    return SIZING_REVERSE_MATRIX[-1]


# Tried all ratios and for the dataset we have correct were ratio, partial ratio, partial token sort ratio, partial token set, token set ratio
# "FM: Ratio", "FM: Partial Ratio", "FM: Partial Sort Ratio", "FM: Token Sort Ratio", "FM: Token Set Ratio"
def fuzzyMatch(issueDesc, linkedIssueDesc):
    fmRatio = fuzz.ratio(issueDesc, linkedIssueDesc)
    fmPartialRatio = fuzz.partial_ratio(issueDesc, linkedIssueDesc)
    fmPartialSortRatio = fuzz.partial_token_sort_ratio(issueDesc, linkedIssueDesc)
    fmTokenSortRatio = fuzz.token_sort_ratio(issueDesc, linkedIssueDesc)
    fmTokenSetRatio = fuzz.token_set_ratio(issueDesc, linkedIssueDesc)
    return fmRatio


def cfMap(jira: JIRA):
    return  # Fetch all fields
    allfields = jira.fields()

    # Make a map from field name -> field id
    return {field["name"]: field["id"] for field in allfields}
    # print(nameMap)


def cfValue(issue, cutomField):
    cfRaw = issue.raw["fields"][cutomField["id"]]
    if cfRaw != None:
        if cutomField["type"] == "array":
            (cf,) = cfRaw
            return cf["value"]
        else:
            return cfRaw["value"]
    return None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python core-summarize.py {component-name}")
        sys.exit(2)
    else:
        exit(main(sys.argv[1:]))
